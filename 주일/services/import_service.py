# -*- coding: utf-8 -*-
"""엑셀/CSV 기존 신우 명단·출석 기록 임포트 서비스.

이전에 엑셀이나 다른 방식으로 관리하던 신우 명단과 출석 기록을
이 프로그램의 DB(users / attendance) 형식으로 변환한다.

템플릿(.xlsx) 구성:
  - '안내' 시트  : 작성 방법
  - '명단' 시트  : 이름|세례|소속|팀|휴대폰|전역일|생일(MMDD)|비고|군종병|이전교회
  - '출석' 시트  : 날짜(YYYY-MM-DD)|예배(일요일/수요일)|이름|소속(선택)|출석시각

동작:
  1) preview  - 파일 파싱 후 요약(신규/기존 매칭/중복/미매칭)과 임시 토큰 반환
  2) commit   - 토큰 기준으로 실제 DB 반영 (사용자 생성 + 출석 추가)
"""
import csv
import datetime
import io
import re
import threading
import time
import uuid

from config import SERVER_ENV
from models.database import db
from models import user as user_model
from models import attendance as attendance_model
from services import report_service
from services.user_service import norm_birthday

# ---------------------------------------------------------------------------
# 템플릿 컬럼 정의
# ---------------------------------------------------------------------------
USER_COLS = ['이름', '세례', '소속', '팀', '휴대폰', '전역일', '생일(MMDD)', '비고', '군종병', '이전교회']
ATTEND_COLS = ['날짜', '예배', '이름', '소속', '출석시각']
GUIDE_LINES = [
    '과거 엑셀이나 수기로 관리하던 신우 명단과 출석 기록을 이 프로그램 DB로 옮기는 템플릿입니다.',
    '',
    '[명단 시트] 신우 한 명이 한 행입니다.',
    '  - 이름      : 필수',
    '  - 세례      : O / X (빈칸 허용)',
    '  - 소속      : 기존 소속과 같게 입력. 없으면 "임시"로 등록됩니다.',
    '  - 팀        : 기존 팀과 같게 입력. 없으면 빈칸.',
    '  - 휴대폰    : 숫자(- 허용)',
    '  - 전역일    : YYYY-MM-DD (예: 2027-03-15)',
    '  - 생일(MMDD): 4자리 (예: 0102)',
    '  - 비고      : 새신우 등 메모. "새신우"라 쓰면 자동으로 새신우 태그 처리.',
    '  - 군종병    : O / X',
    '  - 이전교회  : 이전 교회 이름 (선택)',
    '',
    '[출석 시트] 출석 기록 한 건이 한 행입니다.',
    '  - 날짜      : YYYY-MM-DD (필수. 예: 2026-08-30)',
    '  - 예배      : 일요일 / 수요일 (빈칸이면 현재 모드 기준)',
    '  - 이름      : 명단 시트에 있는 이름 (필수)',
    '  - 소속      : 동명이인 구분용 (선택. 이름이 유일하면 비워도 됨)',
    '  - 출석시각  : HH:MM 또는 HH:MM:SS (선택. 비우면 09:00:00)',
    '',
    '이미 DB에 있는 사용자(이름+소속이 일치)는 새로 만들지 않고 기존 계정에 연결됩니다.',
    '이미 입력된 날짜·예배의 출석 기록은 중복으로 판단해 건너뜁니다.',
]

MODE_ALIASES = {
    '일요일': 'sunday', '일': 'sunday', 'sunday': 'sunday', 'sun': 'sunday',
    '수요일': 'wednesday', '수': 'wednesday', 'wednesday': 'wednesday', 'wed': 'wednesday',
}

_PREVIEWS = {}  # token -> preview result (최근 10개만 유지, 접근순 LRU)
_PREVIEWS_LOCK = threading.Lock()
_PREVIEWS_MAX = 10


def _clean(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _norm_date(v):
    """YYYY-MM-DD 형식으로 통일. 실패 시 ''."""
    s = _clean(v)
    if not s:
        return ''
    s = s.split(' ')[0].split('T')[0]
    s = re.sub(r'[./\\]', '-', s)
    m = re.fullmatch(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            return ''
    if re.fullmatch(r'\d{8}', s):
        try:
            return datetime.date(int(s[:4]), int(s[4:6]), int(s[6:8])).isoformat()
        except ValueError:
            return ''
    return ''


def _norm_time(v):
    s = _clean(v)
    if not s:
        return '09:00:00'
    m = re.search(r'(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?', s)
    if not m:
        hh = int(s) if re.fullmatch(r'\d{1,2}', s) else None
        if hh is None or not (0 <= hh <= 23):
            return '09:00:00'
        return '%02d:00:00' % hh
    hh, mm = int(m.group(1)), int(m.group(2))
    ss = int(m.group(3)) if m.group(3) else 0
    if hh > 23 or mm > 59 or ss > 59:
        return '09:00:00'
    return '%02d:%02d:%02d' % (hh, mm, ss)


def _norm_bool(v):
    s = _clean(v).upper()
    return 1 if s in ('O', '1', 'YES', '예', '군종', '군종병') else 0


def _norm_mode(v, default='sunday'):
    s = _clean(v).lower()
    if not s:
        return default
    if '수' in s or 'wed' in s:
        return 'wednesday'
    return 'sunday'


def _iter_cells(values):
    """행 values에서 None/빈 문자열 정리된 값 나열."""
    for v in values:
        yield _clean(v)


def _sheet_rows(ws):
    """엑셀 시트 → (헤더 dict, 데이터 dict 리스트). 빈 행은 생략."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    # 헤더 행 찾기: '이름' 또는 '날짜' 포함
    _header_idx = None
    for i, row in enumerate(rows[:5]):
        flat = [_clean(x) for x in row]
        if any(x in ('이름', '날짜') for x in flat):
            _header_idx = i
            break
    if _header_idx is None:
        return [], []
    header = [_clean(x) for x in rows[_header_idx]]
    data = []
    for row in rows[_header_idx + 1:]:
        vals = [_clean(x) for x in row]
        if not any(vals):
            continue
        data.append(vals)
    return header, data


def parse_file(filename, raw_bytes):
    """업로드 파일 → (명단 dict 목록, 출석 dict 목록, 안내용 텍스트).

    .xlsx는 openpyxl, .csv는 인코딩(utf-8-sig → cp949 순)으로 읽는다.
    """
    ext = (filename or '').lower()
    if not ext.endswith(('.xlsx', '.csv')):
        return None, None, None
    if ext.endswith('.xlsx'):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        sheets = wb.sheetnames
        users_rows = []
        att_rows = []
        # 시트명 용어 여유 허용
        for sn in sheets:
            if '명단' in sn or '사용자' in sn:
                users_rows = list(_sheet_rows(wb[sn])[1])
            elif '출석' in sn:
                att_rows = list(_sheet_rows(wb[sn])[1])
            elif '안내' in sn:
                pass
        wb.close()
    else:
        text = None
        for enc in ('utf-8-sig', 'utf-8', 'cp949'):
            try:
                text = raw_bytes.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if text is None:
            return None, None, None
        users_rows, att_rows = _csv_rows(text)

    users = [_parse_user_row(r) for r in users_rows]
    attends = [_parse_att_row(r) for r in att_rows]
    return users, attends, None


def _csv_rows(text):
    rows = list(csv.reader(io.StringIO(text)))
    hdr = None
    mode = None
    users_out, att_out = [], []
    for row in rows:
        vals = [_clean(x) for x in row]
        if not any(vals):
            continue
        if hdr is None:
            flat = set(vals)
            if '이름' in flat and '세례' in flat:
                hdr, mode = vals, 'u'
            elif '날짜' in flat and '이름' in flat:
                hdr, mode = vals, 'a'
            else:
                continue
            continue
        if mode == 'u':
            users_out.append(vals)
        else:
            att_out.append(vals)
    return users_out, att_out


def _parse_user_row(vals):
    d = {}
    for i, col in enumerate(USER_COLS):
        d[col] = _clean(vals[i]) if i < len(vals) else ''
    return d


def _parse_att_row(vals):
    d = {}
    for i, col in enumerate(ATTEND_COLS):
        d[col] = _clean(vals[i]) if i < len(vals) else ''
    return d


def _match_existing(conn, name, aff=None, team=None):
    """이름(→소속→팀)으로 기존 사용자 정확 매칭. 모호하면 None."""
    rows = conn.execute('SELECT * FROM users WHERE name=?', (name,)).fetchall()
    if not rows:
        return None
    if len(rows) == 1:
        return rows[0]
    if aff:
        m = [r for r in rows if (r['affiliation'] or '') == aff]
        if m:
            m2 = [r for r in m if (r['team'] or '') == team] if team else m
            if len(m2) == 1:
                return m2[0]
            return None if len(m) > 1 else m[0]
    if team:
        m = [r for r in rows if (r['team'] or '') == team]
        if len(m) == 1:
            return m[0]
    return None


def _preview_cache_check(token):
    with _PREVIEWS_LOCK:
        e = _PREVIEWS.get(token)
        if e and e['expires'] <= time.time():
            _PREVIEWS.pop(token, None)  # 만료 즉시 제거 (메모리 정리)
            return None
        if e:
            e['last'] = time.time()  # LRU 갱신
            return e['data']
    return None


def _cache_put(data):
    token = uuid.uuid4().hex
    now = time.time()
    with _PREVIEWS_LOCK:
        _PREVIEWS[token] = {'expires': now + 3600, 'last': now, 'data': data}
        # 크기 상한 도달 시 만료 항목 먼저 제거, 없으면 가장 오래 접근한 항목 제거
        expired = [k for k, v in _PREVIEWS.items() if v['expires'] <= now]
        for k in expired:
            _PREVIEWS.pop(k, None)
        while len(_PREVIEWS) > _PREVIEWS_MAX:
            oldest = min(_PREVIEWS, key=lambda k: _PREVIEWS[k]['last'])
            _PREVIEWS.pop(oldest, None)
    return token


def build_preview(filename, raw_bytes):
    """파일을 파싱해 임포트 예상 결과를 계산한다. 반환: (ok, payload or msg)"""
    users, attends, _ = parse_file(filename, raw_bytes)
    if users is None:
        return False, {'ok': False, 'msg': '지원하지 않는 파일 형식입니다. .xlsx 또는 .csv 파일을 올려주세요.'}
    if not users and not attends:
        return False, {'ok': False, 'msg': '파일에서 데이터를 찾지 못했습니다. 템플릿 안내를 확인하세요.'}

    user_items = []
    user_key_map = {}  # (name, aff) -> user index
    for i, r in enumerate(users):
        name = r['이름']
        if not name:
            continue
        aff = r['소속']
        team = r['팀']
        key = (name, aff)
        if key not in user_key_map:
            user_key_map[key] = len(user_items)
            user_items.append({
                'idx': len(user_items),
                'name': name, 'aff': aff, 'team': team,
                'baptism': 'O' if _clean(r['세례']).upper() == 'O' else ('X' if _clean(r['세례']).upper() == 'X' else ''),
                'phone': _clean(r['휴대폰']),
                'discharge_date': _norm_date(r['전역일']),
                'birthday': norm_birthday(r['생일(MMDD)']),
                'note': _clean(r['비고']),
                'is_chaplain': _norm_bool(r['군종병']),
                'prev_church': _clean(r['이전교회']),
                'action': 'new', 'uid': None,
            })

    # 기존 매칭 판단
    conn = db()
    try:
        for it in user_items:
            ex = _match_existing(conn, it['name'], it['aff'], it['team'])
            if ex:
                it['action'] = 'match'
                it['uid'] = ex['id']
    finally:
        conn.close()

    # 출석 매칭
    att_items = []
    skip_num = 0
    for r in attends:
        date = _norm_date(r['날짜'])
        if not date:
            skip_num += 1
            continue
        mode = _norm_mode(r['예배'])
        name = r['이름']
        if not name:
            skip_num += 1
            continue
        aff = r['소속']
        time_s = _norm_time(r['출석시각'])
        uid = None
        action = 'nosuch'
        ui = user_key_map.get((name, aff))
        if ui is not None:
            it = user_items[ui]
            uid = it['uid']
            action = 'add'
        else:
            ex = None
            conn = db()
            try:
                ex = _match_existing(conn, name, aff or None, None)
            finally:
                conn.close()
            if ex:
                uid = ex['id']
                action = 'add'
        att_items.append({
            'date': date, 'mode': mode, 'name': name, 'aff': aff,
            'time': time_s, 'uid': uid, 'action': action, 'key': (name, aff),
        })
        if action == 'nosuch':
            skip_num += 1

    user_counts = {'new': 0, 'match': 0}
    for it in user_items:
        user_counts[it['action']] = user_counts.get(it['action'], 0) + 1

    # 중복 출석 선별 (기존 DB 기준)
    dup = 0
    conn = db()
    try:
        _names = set(r['name'] for r in conn.execute('SELECT DISTINCT name FROM users').fetchall())
        for at in att_items:
            if at['action'] != 'add':
                continue
            if at['uid'] is None:
                continue
            if attendance_model.get_attendance(conn, at['uid'], at['date'], at['mode'], SERVER_ENV):
                at['action'] = 'dup'
                dup += 1
    finally:
        conn.close()

    data = {
        'filename': filename,
        'env': SERVER_ENV,
        'users': user_items,
        'attendance': att_items,
        'user_counts': user_counts,
        'att_counts': {
            'add': sum(1 for a in att_items if a['action'] == 'add'),
            'dup': dup,
            'nosuch': sum(1 for a in att_items if a['action'] == 'nosuch'),
            'skip_bad': skip_num,
        },
    }
    token = _cache_put(data)
    return True, {'ok': True, 'token': token, 'summary': _summary(data)}


def _summary(data):
    return {
        'users_new': data['user_counts'].get('new', 0),
        'users_match': data['user_counts'].get('match', 0),
        'attendance_add': data['att_counts']['add'],
        'attendance_dup': data['att_counts']['dup'],
        'attendance_nosuch': data['att_counts']['nosuch'],
        'bad_rows': data['att_counts']['skip_bad'],
        'env': data['env'],
    }


def commit(token):
    """미리보기 결과를 실제 DB에 반영한다."""
    data = _preview_cache_check(token)
    if not data:
        return False, {'ok': False, 'msg': '미리보기 세션이 만료되었습니다. 다시 파일을 올려주세요.'}

    conn = db()
    user_uid_map = {}  # key -> uid
    for it in data['users']:
        if it['action'] == 'match':
            user_uid_map[(it['name'], it['aff'])] = it['uid']
            continue
        # new: 실제 생성 (이름+소속 기준으로 기존 여부 최종 재확인)
        ex = _match_existing(conn, it['name'], it['aff'], it['team'])
        if ex:
            it['uid'] = ex['id']
            it['action'] = 'match'
        else:
            note = it['note']
            if '새신우' in note:
                from services.user_service import normalize_newbie_note
                note = normalize_newbie_note(note)
            new_id = user_model.create_user(conn, {
                'name': it['name'], 'baptism': it['baptism'], 'affiliation': it['aff'],
                'team': it['team'], 'phone': it['phone'], 'discharge_date': it['discharge_date'],
                'birthday': it['birthday'], 'note': note, 'is_chaplain': it['is_chaplain'],
            })
            it['uid'] = new_id
            if it['prev_church']:
                conn.execute('UPDATE users SET prev_church=? WHERE id=?',
                             (it['prev_church'], new_id))
        user_uid_map[(it['name'], it['aff'])] = it['uid']

    added = duped = nosuch = 0
    seen_keys = set()
    for at in data['attendance']:
        uid = at.get('uid')
        if uid is None:
            key = at['key']
            uid = user_uid_map.get(key)
        if uid is None:
            # 명단에 없고 기존에도 없는 이름
            if at['action'] != 'dup':
                nosuch += 1
            continue
        exists = attendance_model.get_attendance(conn, uid, at['date'], at['mode'], SERVER_ENV)
        if exists:
            duped += 1
            continue
        dedupe_key = (uid, at['date'], at['mode'])
        if dedupe_key in seen_keys:
            duped += 1
            continue
        seen_keys.add(dedupe_key)
        user = user_model.get_user(conn, uid)
        if not user:
            nosuch += 1
            continue
        attendance_model.create_attendance(
            conn, uid, user['name'], user['affiliation'],
            at['date'], at['time'], at['mode'], SERVER_ENV)
        added += 1

    _is_new = sum(1 for it in data['users'] if it['action'] == 'new')
    _is_match = sum(1 for it in data['users'] if it['action'] == 'match')
    _new_teams = list({it['team'] for it in data['users'] if it['team']})
    try:
        conn.commit()
    finally:
        conn.close()

    # 보고서 갱신 (양쪽 모드)
    for m in ('sunday', 'wednesday'):
        try:
            report_service.safe_refresh(m, SERVER_ENV)
        except Exception:
            pass

    return True, {
        'ok': True,
        'users_new': _is_new,
        'users_match': _is_match,
        'attendance_add': added,
        'attendance_dup': duped,
        'attendance_nosuch': nosuch,
    }


def build_template_bytes():
    """다운로드용 .xlsx 템플릿 바이트를 생성한다."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    wb.remove(wb.active)

    # 안내 시트
    guide = wb.create_sheet('안내')
    guide.column_dimensions['A'].width = 90
    for i, line in enumerate(GUIDE_LINES, start=1):
        cell = guide.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(vertical='center')
        if line and not line.startswith((' ', '  ')) and not line == '':
            cell.font = Font(bold=True)

    # 명단 시트
    ws = wb.create_sheet('명단')
    head_fill = PatternFill('solid', fgColor='D6E2FF')
    head_font = Font(bold=True, color='1F2937')
    for c, col in enumerate(USER_COLS, start=1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal='center')
    ws.append(['홍길동', 'O', '예비역', '기쁨팀', '010-1234-5678', '', '0102', '새신우', 'X', '서울중앙교회'])
    ws.append(['김이순', 'X', '임시', '', '010-0000-0000', '2027-03-15', '0504', '', '', ''])
    for c in range(1, len(USER_COLS) + 1):
        ws.column_dimensions[chr(64 + c)].width = 14
    ws.column_dimensions['A'].width = 10

    # 출석 시트
    ws2 = wb.create_sheet('출석')
    for c, col in enumerate(ATTEND_COLS, start=1):
        cell = ws2.cell(row=1, column=c, value=col)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal='center')
    ws2.append(['2026-08-30', '일요일', '홍길동', '예비역', '09:31'])
    ws2.append(['2026-08-30', '', '김이순', '', ''])
    for c in range(1, len(ATTEND_COLS) + 1):
        ws2.column_dimensions[chr(64 + c)].width = 14
    ws2.column_dimensions['A'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()